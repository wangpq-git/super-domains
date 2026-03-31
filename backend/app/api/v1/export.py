import io
import csv
import logging
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_db
from app.models.domain import Domain
from app.models.dns_record import DnsRecord
from app.models.platform_account import PlatformAccount

logger = logging.getLogger(__name__)
router = APIRouter()

DOMAIN_CSV_COLUMNS = [
    "domain_name", "platform", "account", "status",
    "expiry_date", "auto_renew", "nameservers", "registration_date",
]

DOMAIN_CSV_HEADERS = [
    "域名", "平台", "账户", "状态",
    "到期日期", "自动续费", "Nameservers", "注册日期",
]

DNS_CSV_COLUMNS = ["record_type", "name", "content", "ttl", "priority", "proxied"]
DNS_CSV_HEADERS = ["类型", "名称", "内容", "TTL", "优先级", "代理"]


async def _get_all_domains(
    db: AsyncSession,
    platform: Optional[str] = None,
    status_filter: Optional[str] = None,
    search: Optional[str] = None,
):
    from sqlalchemy import select
    query = (
        select(
            Domain.id,
            Domain.domain_name,
            Domain.status,
            Domain.registration_date,
            Domain.expiry_date,
            Domain.auto_renew,
            Domain.nameservers,
            PlatformAccount.platform,
            PlatformAccount.account_name,
        )
        .outerjoin(PlatformAccount, PlatformAccount.id == Domain.account_id)
        .where(Domain.status != "removed")
    )
    if platform:
        query = query.where(PlatformAccount.platform == platform)
    if status_filter:
        query = query.where(Domain.status == status_filter)
    if search:
        query = query.where(Domain.domain_name.ilike(f"%{search}%"))
    query = query.order_by(Domain.expiry_date.asc())
    result = await db.execute(query)
    return result.all()


@router.get("/domains")
async def export_domains(
    format: str = Query(default="csv"),
    platform: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    search: Optional[str] = Query(default=None),
    db: AsyncSession = Depends(get_db),
):
    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Unsupported format, use csv or xlsx")

    rows = await _get_all_domains(db, platform=platform, status_filter=status, search=search)

    if format == "csv":
        return await _export_domains_csv(rows)
    else:
        return await _export_domains_xlsx(rows)


async def _export_domains_csv(rows):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(DOMAIN_CSV_HEADERS)

    for row in rows:
        ns = ", ".join(row.nameservers) if row.nameservers else ""
        writer.writerow([
            row.domain_name or "",
            row.platform or "",
            row.account_name or "",
            row.status or "",
            row.expiry_date.strftime("%Y-%m-%d") if row.expiry_date else "",
            "是" if row.auto_renew else "否",
            ns,
            row.registration_date.strftime("%Y-%m-%d") if row.registration_date else "",
        ])

    buffer.seek(0)
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": "attachment; filename=domains.csv"},
    )


async def _export_domains_xlsx(rows):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "域名列表"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(DOMAIN_CSV_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        ns = ", ".join(row.nameservers) if row.nameservers else ""
        values = [
            row.domain_name or "",
            row.platform or "",
            row.account_name or "",
            row.status or "",
            row.expiry_date.strftime("%Y-%m-%d") if row.expiry_date else "",
            "是" if row.auto_renew else "否",
            ns,
            row.registration_date.strftime("%Y-%m-%d") if row.registration_date else "",
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(DOMAIN_CSV_HEADERS) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=domains.xlsx"},
    )


@router.get("/dns/{domain_id}")
async def export_dns(
    domain_id: int,
    format: str = Query(default="csv"),
    db: AsyncSession = Depends(get_db),
):
    from sqlalchemy import select

    if format not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Unsupported format, use csv or xlsx")

    domain_result = await db.execute(select(Domain).where(Domain.id == domain_id))
    domain = domain_result.scalar_one_or_none()
    if not domain:
        raise HTTPException(status_code=404, detail="域名不存在")

    records_result = await db.execute(
        select(DnsRecord)
        .where(DnsRecord.domain_id == domain_id)
        .order_by(DnsRecord.record_type, DnsRecord.name)
    )
    records = records_result.scalars().all()

    if format == "csv":
        return _export_dns_csv(domain, records)
    else:
        return _export_dns_xlsx(domain, records)


def _export_dns_csv(domain, records):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(DNS_CSV_HEADERS)

    for rec in records:
        writer.writerow([
            rec.record_type,
            rec.name,
            rec.content,
            rec.ttl,
            rec.priority or "",
            "是" if rec.proxied else "否",
        ])

    buffer.seek(0)
    filename = f"dns_{domain.domain_name}.csv"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


def _export_dns_xlsx(domain, records):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = f"DNS - {domain.domain_name}"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(DNS_CSV_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, rec in enumerate(records, 2):
        values = [rec.record_type, rec.name, rec.content, rec.ttl, rec.priority or "", "是" if rec.proxied else "否"]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(DNS_CSV_HEADERS) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"dns_{domain.domain_name}.xlsx"
    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
