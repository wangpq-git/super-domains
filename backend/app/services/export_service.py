import io
import csv
import logging
from typing import Optional

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.domain import Domain
from app.models.dns_record import DnsRecord
from app.models.platform_account import PlatformAccount

logger = logging.getLogger(__name__)

DOMAIN_CSV_HEADERS = ["域名", "平台", "账户", "状态", "到期日期", "自动续费", "Nameservers", "注册日期"]
DNS_CSV_HEADERS = ["类型", "名称", "内容", "TTL", "优先级", "代理"]


async def get_all_domains(
    db: AsyncSession,
    *,
    platform: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
):
    query = (
        select(
            Domain.id,
            Domain.domain_name,
            Domain.status,
            Domain.registration_date,
            Domain.expiry_date,
            Domain.auto_renew,
            Domain.nameservers,
            PlatformAccount.platform,
            PlatformAccount.account_name,
        )
        .outerjoin(PlatformAccount, PlatformAccount.id == Domain.account_id)
        .where(Domain.status != "removed")
    )
    if platform:
        query = query.where(PlatformAccount.platform == platform)
    if status:
        query = query.where(Domain.status == status)
    if search:
        query = query.where(Domain.domain_name.ilike(f"%{search}%"))
    query = query.order_by(Domain.expiry_date.asc())
    result = await db.execute(query)
    return result.all()


async def get_dns_records(db: AsyncSession, domain_id: int):
    domain_result = await db.execute(select(Domain).where(Domain.id == domain_id))
    domain = domain_result.scalar_one_or_none()
    if not domain:
        raise ValueError("域名不存在")

    records_result = await db.execute(
        select(DnsRecord)
        .where(DnsRecord.domain_id == domain_id)
        .order_by(DnsRecord.record_type, DnsRecord.name)
    )
    return domain, list(records_result.scalars().all())


def export_domains_csv(rows) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(DOMAIN_CSV_HEADERS)
    for row in rows:
        ns = ", ".join(row.nameservers) if row.nameservers else ""
        writer.writerow([
            row.domain_name or "",
            row.platform or "",
            row.account_name or "",
            row.status or "",
            row.expiry_date.strftime("%Y-%m-%d") if row.expiry_date else "",
            "是" if row.auto_renew else "否",
            ns,
            row.registration_date.strftime("%Y-%m-%d") if row.registration_date else "",
        ])
    return buffer.getvalue().encode("utf-8-sig")


def export_domains_xlsx(rows) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "域名列表"

    header_font = Font(bold=True)
    for col_idx, header in enumerate(DOMAIN_CSV_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, 2):
        ns = ", ".join(row.nameservers) if row.nameservers else ""
        values = [
            row.domain_name or "",
            row.platform or "",
            row.account_name or "",
            row.status or "",
            row.expiry_date.strftime("%Y-%m-%d") if row.expiry_date else "",
            "是" if row.auto_renew else "否",
            ns,
            row.registration_date.strftime("%Y-%m-%d") if row.registration_date else "",
        ]
        for col_idx, val in enumerate(values, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    for col_idx in range(1, len(DOMAIN_CSV_HEADERS) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_dns_csv(domain, records) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(DNS_CSV_HEADERS)
    for rec in records:
        writer.writerow([
            rec.record_type,
            rec.name,
            rec.content,
            rec.ttl,
            rec.priority or "",
            "是" if rec.proxied else "否",
        ])
    return buffer.getvalue().encode("utf-8-sig")
